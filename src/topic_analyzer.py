import pandas as pd
from collections import Counter
import os
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config.base_config import base_config

class TopicAnalyzer:
    def __init__(self, excel_path=None):
        """初始化分析器"""
        self.excel_path = excel_path or base_config.EXCEL_PATH
        self.data_dir = base_config.DATA_DIR
        
    def analyze_topics(self):
        """分析话题"""
        try:
            # 读取Excel文件
            df = pd.read_excel(self.excel_path)
            
            # 使用"关键词"列进行分析
            if '关键词' not in df.columns:
                print("错误：Excel文件中没有'关键词'列！")
                print("现有的列名：", df.columns.tolist())
                return
            
            # 创建话题统计字典
            topic_stats = {}
            
            # 收集所有话题及其互动量
            for _, row in df.iterrows():
                if pd.isna(row['关键词']):
                    continue
                    
                topics = str(row['关键词']).split(',')
                engagement = row['互动量']
                
                for topic in topics:
                    if topic not in topic_stats:
                        topic_stats[topic] = {
                            'count': 0,
                            'total_engagement': 0,
                            'max_engagement': 0,
                            'min_engagement': float('inf')
                        }
                    stats = topic_stats[topic]
                    stats['count'] += 1
                    stats['total_engagement'] += engagement
                    stats['max_engagement'] = max(stats['max_engagement'], engagement)
                    stats['min_engagement'] = min(stats['min_engagement'], engagement)
            
            # 计算每个话题的统计数据
            topic_analysis = []
            for topic, stats in topic_stats.items():
                avg_engagement = stats['total_engagement'] / stats['count']
                topic_analysis.append({
                    'topic': topic,
                    'count': stats['count'],
                    'total_engagement': stats['total_engagement'],
                    'avg_engagement': avg_engagement,
                    'max_engagement': stats['max_engagement'],
                    'min_engagement': stats['min_engagement']
                })
            
            # 按出现次数排序，使用配置的TOP_N值
            topic_analysis.sort(key=lambda x: x['count'], reverse=True)
            top_topics = topic_analysis[:base_config.TOP_N_TOPICS]
            
            # 从文件名提取关键词
            filename = os.path.basename(self.excel_path)
            keyword_match = re.search(r'xiaohongshu_(.+?)_', filename)
            keyword = keyword_match.group(1) if keyword_match else '未知关键词'
            
            # 生成输出文件名
            current_time = datetime.now().strftime('%Y%m%d_%H点%M分')
            output_filename = f'{keyword}_top{base_config.TOP_N_TOPICS}话题_{current_time}.xlsx'
            output_path = os.path.join(self.data_dir, output_filename)
            
            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'Top {base_config.TOP_N_TOPICS} 话题'
            
            # 写入表头
            for col, header in enumerate(base_config.TOPIC_HEADERS, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, topic_data in enumerate(top_topics, 2):
                # 话题列
                cell = ws.cell(row=row_idx, column=1)
                cell.value = topic_data['topic']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 出现次数列
                cell = ws.cell(row=row_idx, column=2)
                cell.value = topic_data['count']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 总互动量列
                cell = ws.cell(row=row_idx, column=3)
                cell.value = int(topic_data['total_engagement'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 平均互动量列（取整）
                cell = ws.cell(row=row_idx, column=4)
                cell.value = int(topic_data['avg_engagement'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 最高互动量列
                cell = ws.cell(row=row_idx, column=5)
                cell.value = int(topic_data['max_engagement'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 最低互动量列
                cell = ws.cell(row=row_idx, column=6)
                cell.value = int(topic_data['min_engagement'])
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            for col, width in base_config.TOPIC_COLUMN_WIDTHS.items():
                ws.column_dimensions[col].width = width
            
            # 保存Excel文件
            wb.save(output_path)
            
            print(f"\n分析完成！")
            print(f"共发现 {len(topic_analysis)} 个话题标签")
            print(f"结果已保存至：{output_filename}")
            
            return top_topics
            
        except Exception as e:
            print(f"分析过程出错: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return None

def main():
    """主函数"""
    analyzer = TopicAnalyzer()
    analyzer.analyze_topics()

if __name__ == '__main__':
    main()