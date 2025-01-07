import pandas as pd
import re
from collections import Counter
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

class ViralTitleAnalyzer:
    def __init__(self, df, viral_threshold=1000):
        """初始化爆款分析器"""
        self.df = df
        self.viral_threshold = viral_threshold
        self.data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
        os.makedirs(self.data_dir, exist_ok=True)

    def analyze_title_structure(self, title):
        """分析标题结构特征"""
        patterns = {
            '数字开头': r'^\d+',
            '感叹句': r'[！!]+',
            '疑问句': r'[？\?]+',
            '表情符号': r'\[.*?\]',  # 小红书特有的表情
            '引用语': r'[「""』」]+',
            '对比句': r'[VS|vs|v\.s\.|比]',
            '省略号': r'\.{3,}|。{3,}|…',
            '强调词': r'绝了|必看|首发|安利|推荐|种草|测评',
            '情感词': r'喜欢|爱了|好看|舒服|完美|超级|真的'
        }
        
        features = []
        for name, pattern in patterns.items():
            if re.search(pattern, title):
                features.append(name)
                
        return features

    def extract_title_templates(self):
        """提取爆款标题模板"""
        viral_posts = self.df[self.df['点赞数'] >= self.viral_threshold].copy()
        
        if viral_posts.empty:
            print(f"没有找到爆款内容（点赞数>={self.viral_threshold}）")
            return None
            
        templates = []
        for _, row in viral_posts.iterrows():
            # 添加数据验证
            if pd.isna(row['笔记标题']):  # 检查标题是否为空
                continue
                
            title = str(row['笔记标题'])  # 确保标题是字符串
            likes = row['点赞数']
            
            template = {
                '原标题': title,
                '点赞数': likes,
                '字数': len(title),
                '结构特征': self.analyze_title_structure(title),
            }
            templates.append(template)
            
        if not templates:  # 如果没有有效的模板
            print("没有找到有效的爆款内容")
            return None
            
        return templates

    def generate_template_report(self, templates, keyword):
        """生成模板库报告"""
        if not templates:
            print("没有发现爆款内容，无法生成报告")
            return
            
        current_time = datetime.now().strftime('%Y%m%d_%H点%M分')
        filename = f'爆款标题模板库_{keyword}_{current_time}.xlsx'
        save_path = os.path.join(self.data_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '爆款标题模板分析'
        
        # 1. 基础统计
        ws['A1'] = '爆款标题模板分析报告'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = f'分析标准：点赞数 >= {self.viral_threshold}'
        ws['A4'] = f'爆款数量：{len(templates)}篇'
        ws['A5'] = f'平均字数：{sum(t["字数"] for t in templates)/len(templates):.1f}字'
        
        # 2. 结构特征统计
        ws['A7'] = '常见结构特征：'
        ws['A7'].font = Font(bold=True)
        all_features = [f for t in templates for f in t['结构特征']]
        feature_stats = Counter(all_features).most_common()
        for i, (feature, count) in enumerate(feature_stats, 1):
            ws[f'A{7+i}'] = f'{feature}: {count}次 ({count/len(templates)*100:.1f}%)'
            
        # 3. 爆款案例展示
        ws['A15'] = 'TOP 10爆款标题示例：'
        ws['A15'].font = Font(bold=True)
        sorted_templates = sorted(templates, key=lambda x: x['点赞数'], reverse=True)[:10]
        for i, template in enumerate(sorted_templates, 1):
            ws[f'A{15+i}'] = template['原标题']
            ws[f'B{15+i}'] = f"点赞数: {template['点赞数']}"
            ws[f'C{15+i}'] = f"特征: {', '.join(template['结构特征'])}"
            
        # 4. 标题模板建议
        ws['A30'] = '标题模板建议：'
        ws['A30'].font = Font(bold=True)
        template_suggestions = [
            "数字 + 核心关键词 + 强调词（例：3个让鞋子绝美的小技巧）",
            "疑问句 + 解决方案（例：冬天鞋子总是湿？这个方法超管用）",
            "对比句式（例：平价VS大牌，这双鞋性价比太高了）",
            "强调词 + 核心卖点（例：必看！让鞋子持久如新的秘密）",
            "场景化描述（例：下雨天也能美美的，防水鞋首选）",
            "情感共鸣 + 产品优势（例：真的被这双鞋子惊艳到了）",
            "测评 + 核心卖点（例：超详细测评！这双鞋子防水效果绝了）"
        ]
        for i, suggestion in enumerate(template_suggestions, 1):
            ws[f'A{30+i}'] = f"{i}. {suggestion}"
            
        # 设置格式
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        
        # 保存文件
        wb.save(save_path)
        print(f"\n爆款标题模板库已保存: {filename}")

# 如果需要独立运行，可以添加以下代码
if __name__ == '__main__':
    try:
        # 读取Excel文件
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                                'data', 
                                "xiaohongshu_鞋子_20241223_16点04分 403条.xlsx")
        
        if not os.path.exists(excel_path):
            print(f"错误：找不到文件")
            exit()
            
        df = pd.read_excel(excel_path)
        print(f"成功加载数据，共 {len(df)} 条记录")
        
        # 创建分析器并执行分析
        viral_analyzer = ViralTitleAnalyzer(df)
        templates = viral_analyzer.extract_title_templates()
        if templates:
            viral_analyzer.generate_template_report(templates, "鞋子")
            
    except Exception as e:
        print(f"分析过程出错: {str(e)}")
        import traceback
        print(traceback.format_exc())