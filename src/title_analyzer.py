import pandas as pd
from collections import Counter
import os
from datetime import datetime
import re
import jieba
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import matplotlib
from config.base_config import base_config

matplotlib.use('Agg')  # 使用非交互式后端

class TitleAnalyzer:
    def __init__(self, df):
        """初始化分析器"""
        self.df = df
        self.data_dir = base_config.DATA_DIR
        
    def extract_common_phrases(self, titles):
        """提取标题中的共性词组"""
        # 使用配置中的关键词列表
        relevant_keywords = base_config.TITLE_RELEVANT_KEYWORDS
        
        # 分词并统计词频
        word_freq = Counter()
        phrase_patterns = []
        
        # 第一步：对所有标题进行分词
        all_words_in_titles = []
        for title in titles:
            # 使用jieba的搜索引擎模式，更容易识别短语
            words = [w for w in jieba.cut_for_search(title) if len(w) > 1]
            all_words_in_titles.append(words)
            word_freq.update(words)
        
        # 第二步：寻找共同出现的词组
        common_phrases = []
        for i in range(len(all_words_in_titles)):
            words = all_words_in_titles[i]
            
            # 在单个标题内寻找可能的词组
            for j in range(len(words)-1):
                for k in range(j+1, min(j+4, len(words))):  # 最多看相邻的3个词
                    phrase_words = words[j:k+1]
                    if len(phrase_words) >= 2:  # 至少2个词
                        phrase = ''.join(phrase_words)
                        # 检查这个词组在其他标题中的出现次数
                        count = sum(1 for title in titles if phrase in title)
                        if count >= 2:  # 至少在2个标题中出现
                            common_phrases.append({
                                'phrase': phrase,
                                'frequency': count,
                                'words': phrase_words,
                                'example_titles': [t for t in titles if phrase in t][:3]  # 保存最多3个示例标题
                            })
        
        # 第三步：去重和排序
        seen_phrases = set()
        unique_phrases = []
        for p in common_phrases:
            if p['phrase'] not in seen_phrases:
                seen_phrases.add(p['phrase'])
                unique_phrases.append(p)
        
        # 按频率和词组长度排序（优先选择频率高且较长的词组）
        unique_phrases.sort(key=lambda x: (x['frequency'], len(x['phrase'])), reverse=True)
        
        # 第四步：过滤掉不够相关的词组
        filtered_phrases = []
        for p in unique_phrases:
            # 检查词组是否有意义
            if len(p['words']) >= 2 and (  # 至少2个词
                # 条件1：包含相关关键词
                any(w in relevant_keywords for w in p['words']) or
                # 条件2：高频出现（出现3次以上）且词组长度合适
                (p['frequency'] >= 3 and len(p['phrase']) >= 3) or
                # 条件3：特别高频（出现5次以上）
                p['frequency'] >= 5
            ):
                filtered_phrases.append(p)
        
        return filtered_phrases[:10]  # 返回top 10共性词组

    def analyze_similar_titles(self):
        """分析相似标题"""
        print("\n=== 开始标题分析 ===")
        title_analysis = []
        
        # 1. 提取所有标题的关键词
        title_info = {}  # 存储标题的关键词和点赞数
        for idx, row in self.df.iterrows():
            title = row['笔记标题']
            likes = row.get('点赞数', 0)  # 获取点赞数
            
            # 确保标题是字符串类型
            if pd.isna(title):  # 处理空值
                continue
            title = str(title)  # 转换为字符串
            
            # 清理标题
            clean_title = re.sub(r'[^\w\s]', '', title).strip()
            # 分词
            keywords = [word for word in jieba.cut(clean_title) if len(word) > 1]
            title_info[title] = {
                'keywords': keywords,
                'likes': likes
            }
        
        # 2. 统计关键词出现次数
        all_keywords = []
        for info in title_info.values():
            all_keywords.extend(info['keywords'])
        keyword_counter = Counter(all_keywords)
        
        # 3. 按关键词对标题进行分组
        keyword_titles = {}
        for title, info in title_info.items():
            if info['keywords']:
                main_keyword = max(info['keywords'], key=lambda x: keyword_counter[x])
                if main_keyword not in keyword_titles:
                    keyword_titles[main_keyword] = []
                keyword_titles[main_keyword].append((title, info['likes']))
        
        # 4. 整理分析结果
        for keyword, titles in keyword_titles.items():
            # 按点赞数降序排序标题
            sorted_titles = sorted(titles, key=lambda x: x[1], reverse=True)
            title_analysis.append({
                '关键词': keyword,
                '出现次数': len(titles),
                '相关标题': [t[0] for t in sorted_titles],  # 只保留标题，不包含点赞数
                '标题数量': len(titles)
            })
        
        # 按出现次数降序排序关键词
        title_analysis.sort(key=lambda x: x['出现次数'], reverse=True)
        return title_analysis

    def highlight_keyword(self, title, keyword):
        """在标题中高亮关键词"""
        return title.replace(keyword, f'【{keyword}】')

    def save_analysis_results(self, analysis_results, keyword):
        """保存分析结果"""
        current_time = datetime.now().strftime('%Y%m%d_%H点%M分')
        filename = f'标题分析_{keyword}_{current_time}.xlsx'
        save_path = os.path.join(self.data_dir, filename)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = '标题分析'
        
        # 直接写入数据到DataFrame
        df_results = pd.DataFrame(analysis_results)
        df_results = df_results[['关键词', '出现次数', '标题数量', '相关标题']]
        
        # 写入表头 - 修改为深灰色背景
        headers = ['关键词', '出现次数', '标题数量', '相关标题']
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # 写入数据
        for row_idx, row in df_results.iterrows():
            # 关键词列
            cell = ws_data.cell(row=row_idx+2, column=1)
            cell.value = row['关键词']
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 出现次数列
            cell = ws_data.cell(row=row_idx+2, column=2)
            cell.value = row['出现次数']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 标题数量列
            cell = ws_data.cell(row=row_idx+2, column=3)
            cell.value = row['标题数量']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 相关标题列
            highlighted_titles = []
            for i, title in enumerate(row['相关标题'], 1):
                highlighted_title = self.highlight_keyword(title, row['关键词'])
                highlighted_titles.append(f"{i}. {highlighted_title}")
            
            cell = ws_data.cell(row=row_idx+2, column=4)
            cell.value = '\n'.join(highlighted_titles)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 设置列宽和格式
        ws_data.column_dimensions['A'].width = 15
        ws_data.column_dimensions['B'].width = 12
        ws_data.column_dimensions['C'].width = 12
        ws_data.column_dimensions['D'].width = 100
        
        # 设置行高
        ws_data.row_dimensions[1].height = 25  # 表头行高
        
        wb.save(save_path)
        print(f"\n分析结果已保存至：{filename}")

    def analyze_viral_patterns(self):
        """分析爆款标题模板"""
        # 使用配置中的爆款阈值
        VIRAL_THRESHOLD = base_config.TITLE_VIRAL_THRESHOLD
        
        # 收集所有爆款标题
        viral_titles = []
        for _, row in self.df.iterrows():
            title = row['笔记标题']
            likes = row.get('点赞数', 0)
            
            if pd.isna(title) or pd.isna(likes):
                continue
                
            title = str(title)
            likes = float(likes) if not isinstance(likes, (int, float)) else likes
            
            if likes >= VIRAL_THRESHOLD:
                viral_titles.append({
                    'title': title,
                    'likes': likes
                })
        
        # 提取所有爆款标题的共性词
        all_viral_titles = [post['title'] for post in viral_titles]
        common_phrases = self.extract_common_phrases(all_viral_titles)
        
        # 使用配置中的模板
        viral_patterns = base_config.TITLE_PATTERNS
        
        template_stats = {
            category: {
                'count': 0,
                'avg_likes': 0,
                'examples': [],
            } for category in viral_patterns
        }
        
        # 分析每个标题的模板类型
        for post in viral_titles:
            title = post['title']
            likes = post['likes']
            
            for category, patterns in viral_patterns.items():
                for pattern in patterns:
                    if re.search(pattern, title):
                        template_stats[category]['count'] += 1
                        template_stats[category]['examples'].append({
                            'title': title,
                            'likes': likes,
                            'matched_pattern': pattern
                        })
                        break

        # 计算统计信息
        for category in template_stats:
            examples = template_stats[category]['examples']
            if examples:
                template_stats[category]['avg_likes'] = sum(e['likes'] for e in examples) / len(examples)
                template_stats[category]['examples'] = sorted(
                    examples, 
                    key=lambda x: x['likes'], 
                    reverse=True
                )[:5]
        
        template_library = self._generate_template_library(template_stats)
        return template_stats, template_library, common_phrases

    def analyze_low_follower_viral(self):
        """分析低粉爆款笔记"""
        MIN_LIKES = base_config.TITLE_MIN_LIKES
        
        viral_posts = []
        for _, row in self.df.iterrows():
            title = row['笔记标题']
            likes = row.get('点赞数', 0)
            followers = row.get('粉丝数', 0)
            
            if pd.isna(title) or pd.isna(likes) or pd.isna(followers):
                continue
            
            likes = float(likes) if not isinstance(likes, (int, float)) else likes
            followers = float(followers) if not isinstance(followers, (int, float)) else followers
            
            if likes > followers and likes > MIN_LIKES:
                viral_posts.append({
                    'title': str(title),
                    'likes': likes,
                    'followers': followers,
                    'ratio': likes/followers if followers > 0 else float('inf'),
                    'row_data': row
                })
        
        viral_posts.sort(key=lambda x: x['likes'], reverse=True)
        return viral_posts

    def save_low_follower_viral_results(self, viral_posts, keyword):
        """保存低粉爆款分析结果"""
        if not viral_posts:
            print("\n未发现低粉爆款笔记")
            return
            
        current_time = datetime.now().strftime('%Y%m%d_%H点%M分')
        count = len(viral_posts)
        filename = f'xiaohongshu_{keyword}_{current_time} {count}条.xlsx'
        save_path = os.path.join(self.data_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '低粉爆款分析'
        
        # 添加判断标准说明
        ws['A1'] = f"判断标准：点赞数>粉丝数 且 点赞数>{base_config.TITLE_MIN_LIKES}"
        ws['A1'].font = Font(bold=True)
        ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(len(self.df.columns))}1')
        
        # 使用原始DataFrame的列名作为表头
        headers = self.df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # 写入数据（保持原始数据的所有列）
        for row_idx, post in enumerate(viral_posts, 3):
            row_data = post['row_data']
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical='center')  # 垂直居中
                
                # 如果是标题列，启用自动换行
                if col_idx == 1:  # 假设标题是第一列
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(2, len(viral_posts) + 3):  # 从表头开始，包括数据行
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
        
        wb.save(save_path)
        print(f"\n发现 {count} 条低粉爆款笔记")
        print(f"低粉爆款分析已保存至：{filename}")

    def save_viral_templates(self, template_stats, template_library, common_phrases, keyword):
        """保存爆款标题模板分析结果"""
        current_time = datetime.now().strftime('%Y%m%d_%H点%M分')
        filename = f'爆款标题模板_{keyword}_{current_time}.xlsx'
        save_path = os.path.join(self.data_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # 1. 创建共性词分析页
        ws_common = wb.active
        ws_common.title = '爆款共性词分析'
        
        # 添加判断标准说明
        ws_common['A1'] = f"判断标准：点赞数 >= {base_config.TITLE_VIRAL_THRESHOLD}"
        ws_common['A1'].font = Font(bold=True)
        ws_common.merge_cells('A1:C1')
        
        # 写入共性词表头
        headers = ['共性词组', '出现频率', '示例标题']
        for col, header in enumerate(headers, 1):
            cell = ws_common.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        # 写入共性词数据
        row = 3
        for phrase in common_phrases:
            ws_common.cell(row=row, column=1, value=phrase['phrase'])
            ws_common.cell(row=row, column=2, value=phrase['frequency'])
            examples = '\n'.join([f"{i+1}. {title}" for i, title in enumerate(phrase['example_titles'])])
            ws_common.cell(row=row, column=3, value=examples)
            row += 1
        
        # 设置列宽
        ws_common.column_dimensions['A'].width = 20
        ws_common.column_dimensions['B'].width = 12
        ws_common.column_dimensions['C'].width = 60
        
        # 2. 创建模板分析页
        ws_analysis = wb.create_sheet(title='爆款标题分析')
        
        # 添加判断标准说明
        ws_analysis['A1'] = f"判断标准：点赞数 >= {base_config.TITLE_VIRAL_THRESHOLD}"
        ws_analysis['A1'].font = Font(bold=True)
        ws_analysis.merge_cells('A1:D1')
        
        # 写入表头
        headers = ['模板类型', '使用频率', '平均点赞', '成功案例']
        for col, header in enumerate(headers, 1):
            cell = ws_analysis.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        row = 3
        for category, stats in template_stats.items():
            if stats['count'] > 0:
                ws_analysis.cell(row=row, column=1, value=category)
                ws_analysis.cell(row=row, column=2, value=stats['count'])
                ws_analysis.cell(row=row, column=3, value=f"{stats['avg_likes']:.0f}")
                
                examples = '\n'.join([
                    f"{i+1}. {e['title']} (👍{e['likes']})" 
                    for i, e in enumerate(stats['examples'])
                ])
                ws_analysis.cell(row=row, column=4, value=examples)
                row += 1
        
        # 3. 创建模板库页
        ws_templates = wb.create_sheet(title='标题模板库')
        
        # 添加判断标准说明到模板库页
        ws_templates['A1'] = f"判断标准：点赞数 >= {base_config.TITLE_VIRAL_THRESHOLD}"
        ws_templates['A1'].font = Font(bold=True)
        ws_templates.merge_cells('A1:D1')
        
        headers = ['类型', '模板', '示例', '点赞数']
        for col, header in enumerate(headers, 1):
            cell = ws_templates.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
        
        row = 3
        for category, templates in template_library.items():
            for template_info in templates:
                ws_templates.cell(row=row, column=1, value=category)
                ws_templates.cell(row=row, column=2, value=template_info['template'])
                ws_templates.cell(row=row, column=3, value=template_info['original'])
                ws_templates.cell(row=row, column=4, value=template_info['likes'])
                row += 1
        
        # 设置列宽
        ws_analysis.column_dimensions['A'].width = 15
        ws_analysis.column_dimensions['B'].width = 12
        ws_analysis.column_dimensions['C'].width = 12
        ws_analysis.column_dimensions['D'].width = 60
        
        ws_templates.column_dimensions['A'].width = 15
        ws_templates.column_dimensions['B'].width = 30
        ws_templates.column_dimensions['C'].width = 60
        ws_templates.column_dimensions['D'].width = 12
        
        # 设置自动换行和对齐方式
        for ws in [ws_common, ws_analysis, ws_templates]:
            for row in ws.iter_rows(min_row=3):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        wb.save(save_path)
        print(f"\n爆款标题模板分析已保存至：{filename}")

    def _generate_template_library(self, template_stats):
        """根据分析结果生成可复用的标题模板库"""
        template_library = {}
        
        for category, stats in template_stats.items():
            if not stats['examples']:
                continue
                
            templates = []
            for example in stats['examples']:
                title = example['title']
                pattern = example['matched_pattern']
                
                # 提取模板（替换具体词汇为通用占位符）
                template = title
                template = re.sub(r'\d+', '{N}', template)  # 替换具体数字
                template = re.sub(r'[a-zA-Z\u4e00-\u9fff]+鞋', '{产品}', template)  # 替换具体产品
                template = re.sub(r'[a-zA-Z\u4e00-\u9fff]+牌', '{品牌}', template)  # 替换具体品牌
                
                templates.append({
                    'template': template,
                    'original': title,
                    'likes': example['likes']
                })
            
            template_library[category] = templates
        
        return template_library

def main():
    """主函数"""
    try:
        # 指定要分析的文件名
        target_file = base_config.EXCEL_FILENAME
        excel_path = base_config.EXCEL_PATH
        
        if not os.path.exists(excel_path):
            print(f"错误：找不到文件 {target_file}")
            return
        
        print(f"正在分析文件: {target_file}")
        
        # 读取数据
        df = pd.read_excel(excel_path)
        print(f"成功加载数据，共 {len(df)} 条记录")
        
        # 检查必要的列是否存在
        if '笔记标题' not in df.columns:
            print("错误：Excel文件中没有'笔记标题'列！")
            print("现有的列名：", df.columns.tolist())
            return
            
        # 提取关键词
        keyword_match = re.search(r'xiaohongshu_(.+?)_', target_file)
        keyword = keyword_match.group(1) if keyword_match else '未知关键词'
        
        # 创建分析器并执行分析
        analyzer = TitleAnalyzer(df)
        analysis_results = analyzer.analyze_similar_titles()
        analyzer.save_analysis_results(analysis_results, keyword)
        
        # 添加爆款标题分析
        print("\n开始爆款标题模板分析...")
        template_stats, template_library, common_phrases = analyzer.analyze_viral_patterns()
        analyzer.save_viral_templates(template_stats, template_library, common_phrases, keyword)
        
        # 添加低粉爆款分析
        print("\n开始低粉爆款分析...")
        viral_posts = analyzer.analyze_low_follower_viral()
        analyzer.save_low_follower_viral_results(viral_posts, keyword)
        
    except Exception as e:
        print(f"分析过程出错: {str(e)}")
        import traceback
        print(traceback.format_exc())
    
    print("\n=== 分析完成 ===")

if __name__ == '__main__':
    main()